from openpyxl import Workbook, load_workbook
import json
import argparse
import sys
import os
from settings import *
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from dbclass import Doc, Department, Bundle, Base

engine = create_engine('mysql+pymysql://{}:{}@localhost:{}/{}'.format(USER, PASSWORD, PORT, DBNAME) , echo=False)


def getbegin(ws, rownum, col):
    while True:
        # print(ws[f"{col}{rownum}"].value)
        if ws[f"{col}{rownum}"].value == None:
            rownum += 1
        else:
            return rownum


def checkexcel(ws, sheetname):
    for i in range(7, ws.max_row + 1):
        if ws[f"C{i}"].value is None and ws[f"G{i}"].value is not None:
            print("check "+ sheetname + " row: " + str(i) + " " + str(ws[f"G{i}"].value))

def parse(ws, sheetname):
    boxlist = []
    first = True
    for i in range(7, ws.max_row + 1):
        if first:
            begin = i
            first = False
            boxno = ws[f'A{i}'].value
            continue

        if ws[f"A{i}"].value != None:
            # end = getend(i)
            end = i-1
            boxlist.append({"box": boxno, "begin": begin, "end": end})

            begin = getbegin(ws, i, "C")
            boxno = ws[f'A{i}'].value

    boxlist.append({ "box": boxno, "begin": begin, "end": i-1})

    for ke, box in enumerate(boxlist):
        noberkas = ws[f"B{box['begin']}"].value
        first = True
        berkaslist = []
        for i in range(box['begin'], box['end']+1):
            if first:
                begin = box['begin']
                first = False
                berkasno = ws[f'B{i}'].value
                continue

            if ws[f"B{i}"].value != None and  str(ws[f"B{i}"].value).strip() != "":
                end = i-1
                index = ""
                perus = ""
                nourutlist = []
                urutfirst = True
                for i2 in range(begin, end):
                    if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
                        try:
                            index = index + " " + ws[f"E{i2}"].value
                        except:
                            pass

                    if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
                        try:
                            perus = perus + " " + ws[f"F{i2}"].value
                        except:
                            pass
                    
                    if urutfirst:
                        urutfirst = False
                        nourut = int(ws[f"C{i2}"].value)
                        uraian = ws[f"G{i2}"].value
                        jumlah = ws[f"I{i2}"].value
                        # input(nourut)
                    else:
                        if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
                            # nourutlist.append({"nourut": nourut, "uraian": uraian })
                            nourutlist.append({"nourut": nourut, "uraian": uraian, "jumlah": jumlah })
                            # input(nourutlist)
                            nourut = int(ws[f"C{i2}"].value)
                            uraian = ws[f"G{i2}"].value
                            jumlah = ws[f"I{i2}"].value
                        else:
                            # print(box['box'], ws[f"G{i2}"].value, i2)
                            if ws[f"G{i2}"].value is not None and str(ws[f"C{i2}"].value).strip() != "":
                                uraian += "\n" + str(ws[f"G{i2}"].value)

                                # input(uraian)
                            # pass


                            # try:
                            #     nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
                            # except:
                            #     pass
                nourutlist.append({"nourut": nourut, "uraian": uraian, "jumlah": jumlah })

                dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin": begin, "end": end, "data": nourutlist}
                berkaslist.append(dtemp)
                berkasno = ws[f"B{i}"].value
                begin = getbegin(ws, i, "C")

        #DATA BERKAS TERBAWAH
        index = ""
        perus = ""
        nourutlist = []
        urutfirst = True

        for i2 in range(begin, box['end']+1):
            if ws[f"E{i2}"].value != None and str(ws[f"E{i2}"].value).strip() != "":
                try:
                    index = index + " " + ws[f"E{i2}"].value
                except:
                    pass

            if ws[f"F{i2}"].value != None and str(ws[f"F{i2}"].value).strip() != "":
                try:
                    perus = perus + " " + ws[f"F{i2}"].value
                except:
                    pass

            # if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
            #     # try:
            #     nourutlist.append({"nourut": ws[f"C{i2}"].value, "uraian": ws[f"G{i2}"].value, "jumlah": ws[f"I{i2}"].value })
                # except:
                #     pass
            if urutfirst:
                urutfirst = False
                nourut = int(ws[f"C{i2}"].value)
                uraian = ws[f"G{i2}"].value
                jumlah = ws[f"I{i2}"].value
                # input(str(nourut) + "xx")
            else:
                if ws[f"C{i2}"].value != None and str(ws[f"C{i2}"].value).strip() != "":
                    # nourutlist.append({"nourut": nourut, "uraian": uraian })
                    nourutlist.append({"nourut": nourut, "uraian": uraian, "jumlah": jumlah })
                    nourut = int(ws[f"C{i2}"].value)
                    uraian = ws[f"G{i2}"].value
                    jumlah = ws[f"I{i2}"].value
                else:
                    # print(box['box'], ws[f"G{i2}"].value, i2)
                    if ws[f"G{i2}"].value is not None and str(ws[f"C{i2}"].value).strip() != "":
                        uraian += "\n" + str(ws[f"G{i2}"].value)
                        # input(uraian)

        nourutlist.append({"nourut": nourut, "uraian": uraian, "jumlah": jumlah })
        dtemp = {"berkas": berkasno, "kode": ws[f"D{begin}"].value, "tahun": str(ws[f"H{begin}"].value), "ket": str(ws[f"J{begin}"].value), "index": perus.strip() + "\n" + index.strip(), "begin":begin, "end": box['end'], "data": nourutlist}
        berkaslist.append(dtemp)
        boxlist[ke]['data'] = berkaslist
    return boxlist

def listtodb(boxlist, sheetname, session):
    defcode = boxlist[0]['data'][0]['kode']
    dep = Department(name=sheetname, defcode=defcode, link=str(sheetname).replace(' ', "_").lower() )
    session.add(dep)
    session.flush()
    session.commit()
    departmentId = dep.id
    for box in boxlist:
        box_number = box['box']
        for bundle in box['data']:
            bundle_number = bundle['berkas']
            if bundle['ket'] == 'None':
                ket = ''
            else:
                ket = "".join(bundle['ket']).upper()

            if bundle['tahun'] == 'None':
                year = ''
            else:
                year = "".join(bundle['tahun'])
            code = bundle['kode']
            if code is None:
                code = defcode
            bundleinsert = Bundle(department_id=departmentId, box_number=box_number, bundle_number=bundle_number, code=code, title=bundle['index'], year=year, orinot=ket)
            session.add(bundleinsert)
            session.flush()
            for doc in bundle['data']:
                if doc['jumlah']  is None:
                    jumlah = 1
                else:
                    jumlah = doc['jumlah']

                # if doc['uraian'] is not None and doc['nourut'] is None:
                #     print("check:",doc['uraian'], doc['nourut'])
                #     sys.exit()
                if str(jumlah).strip() == "":
                    jumlah = 1
                docinsert = Doc(bundle_id=bundleinsert.id, doc_number=doc['nourut'], doc_count=jumlah, description=doc['uraian'])
                session.add(docinsert)
                print("record number: ", doc['nourut'], "Inserted" )

    session.flush()
    session.commit()


def main():
    # parser = argparse.ArgumentParser(description="Get data from excel and save them to json")
    # parser.add_argument('-input', '--xlsinput', type=str,help="XLSX File Input")
    # parser.add_argument('-sname', '--sheetname', type=str,help="Sheet Name of XLSX file")
    # parser.add_argument('-output', '--jsonoutput', type=str,help="File output in json")

    # args = parser.parse_args()
    # if not (args.xlsinput[-5:] == '.xlsx' or args.xlsinput[-5:] == '.xlsm'):
    #     input('input the right XLSX or XLSM file')
    #     sys.exit()

    # isExist = os.path.exists(args.xlsinput)
    # if not isExist:
    #     input(args.xlsinput + " does not exist")
    #     sys.exit()

    # fname = "Daftar Arsip.xlsx"
    # wb = load_workbook(filename=fname)
    # ws = wb["IRIGASI"]
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    session = Session(engine)

    mainboxlist = []
    wb = load_workbook(filename=EXCEL_FILE)
    for sheetname in EXCEL_SHEET:
        input(sheetname)
        print(sheetname, "generating...", end="", flush=True)
        ws = wb[sheetname]
        # checkexcel(ws, sheetname)
        datalist = parse(ws, sheetname)
        listtodb(boxlist=datalist, sheetname=sheetname, session=session)
        print("Success")
    
    
    # with open(args.jsonoutput, 'w') as file:
    #     json.dump(mainboxlist, file)

if __name__ == '__main__':
    main()
